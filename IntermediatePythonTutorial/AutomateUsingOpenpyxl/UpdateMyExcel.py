############################
# Update Excel file
##################################

import openpyxl as xl
from openpyxl.chart import BarChart,Reference

def update_process(filename,sheetname = "Sheet1"):
        wb = xl.load_workbook(filename)
        sheet = wb[sheetname]

        #cell1  = sheet["a1"]
        #cell = sheet.cell(1,1)

        # print(cell.value)
        # print(sheet.max_row)
        # print(sheet.max_column)

        for row in range(2,sheet.max_row + 1):
            cell = sheet.cell(row,3)
            print(cell.value)
            correct_value = cell.value * 0.9
            correct_value_cell = sheet.cell(row,4)
            correct_value_cell.value = correct_value
            print(correct_value)



        # select cloumn-d to create chart
        values = Reference(sheet,
                min_row = 2, 
                max_row = sheet.max_row, 
                min_col = 4,
                max_col = 4)

        chart = BarChart()
        chart.add_data(values)
        sheet.add_chart(chart,"e2")

        wb.save(filename)

#############################

print("Calling function ...")
update_process("transactions.xlsx")
print("End main.")

################################
#################################
